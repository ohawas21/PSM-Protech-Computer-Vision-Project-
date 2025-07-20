import os
import re
import logging
from pathlib import Path
from io import BytesIO
import warnings
import shutil
import zipfile
import yaml
import camelot
import cv2
import numpy as np
from PIL import Image
from pdf2image import convert_from_bytes
import pytesseract
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger(__name__)

# Suppress Camelot warnings for image-based pages
warnings.filterwarnings('ignore', category=UserWarning)

# Tesseract config: only digits and dot
TESS_CONFIG = "--psm 7 --oem 3 -c tessedit_char_whitelist=0123456789."

class TextExtractor:
    def __init__(self):
        self.temp_dir = Path("temp_processing")
        self.temp_dir.mkdir(parents=True, exist_ok=True)

    def convert_to_pdf(self, image_path: str) -> Path:
        """Convert image to PDF for processing"""
        buf = BytesIO()
        Image.open(image_path).convert("RGB").save(buf, format="PDF")
        buf.seek(0)
        pdf_path = self.temp_dir / f"{Path(image_path).stem}.pdf"
        pdf_path.write_bytes(buf.getvalue())
        return pdf_path

    def try_camelot(self, pdf_path: Path) -> str:
        """Try to extract text using Camelot"""
        try:
            tables = camelot.read_pdf(str(pdf_path), flavor="stream", pages="1")
            if tables:
                text = "\n".join([table.df.to_string() for table in tables])
                logger.info(f"Successfully extracted text using Camelot from {pdf_path.name}")
                return text
        except Exception as e:
            logger.warning(f"Camelot extraction failed for {pdf_path.name}: {str(e)}")
        return ""

    def render_to_image(self, file_path: str) -> Image.Image:
        """Convert PDF to image or load image directly"""
        try:
            if file_path.lower().endswith('.pdf'):
                pages = convert_from_bytes(Path(file_path).read_bytes(), dpi=300)
                return pages[0] if pages else None
            else:
                return Image.open(file_path)
        except Exception as e:
            logger.error(f"Error converting {file_path}: {str(e)}")
            return None

    def preprocess_image(self, img: Image.Image) -> Image.Image:
        """Preprocess image for better OCR"""
        arr = cv2.cvtColor(np.array(img), cv2.COLOR_RGB2GRAY)
        blur = cv2.GaussianBlur(arr, (3,3), 0)
        _, th = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        inv = cv2.bitwise_not(th)
        up = cv2.resize(inv, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
        return Image.fromarray(up)

    def extract_text(self, file_path: str) -> str:
        """Extract text from file using multiple methods"""
        file_path = str(file_path)
        text = ""
        
        # Try Camelot first for PDFs
        if file_path.lower().endswith('.pdf'):
            text = self.try_camelot(Path(file_path))
            if text.strip():
                return text

        # If Camelot fails or it's an image, try OCR
        img = self.render_to_image(file_path)
        if img:
            preprocessed = self.preprocess_image(img)
            text = pytesseract.image_to_string(preprocessed)
            logger.info(f"Extracted text using OCR from {Path(file_path).name}")
            return text.strip()
            
        return ""

# Global extractor instance
text_extractor = TextExtractor()

def load_config():
    """Load configuration from config.yaml"""
    config_path = os.path.join(os.path.dirname(__file__), '..', 'config.yaml')
    try:
        with open(config_path, 'r') as f:
            return yaml.safe_load(f)
    except Exception as e:
        print(f"Warning: Could not load config: {e}")
        return {}

def create_organized_image_export(input_data, original_pdf_name):
    """Create organized image folders for verification"""
    config = load_config()
    if not config.get('enable_image_export', True):
        return None
    
    export_base_dir = config.get('export_images_dir', 'export_images')
    document_export_dir = os.path.join(export_base_dir, original_pdf_name)
    
    # Clean and create export directory
    if os.path.exists(document_export_dir):
        shutil.rmtree(document_export_dir)
    os.makedirs(document_export_dir, exist_ok=True)
    
    print(f"Creating organized image export in: {document_export_dir}")
    
    # Group by parent images (R1 crops)
    parent_groups = {}
    for item in input_data:
        if not isinstance(item, dict) or 'original_path' not in item:
            continue
        parent_path = item['original_path']
        if parent_path not in parent_groups:
            parent_groups[parent_path] = []
        parent_groups[parent_path].append(item)
    
    exported_folders = []
    
    for parent_path, sub_items in parent_groups.items():
        if not os.path.exists(parent_path):
            continue
            
        # Create folder name from parent image
        parent_name = os.path.splitext(os.path.basename(parent_path))[0]
        parent_folder = os.path.join(document_export_dir, parent_name)
        os.makedirs(parent_folder, exist_ok=True)
        
        # Copy parent image
        parent_dest = os.path.join(parent_folder, f"00_parent_{os.path.basename(parent_path)}")
        shutil.copy2(parent_path, parent_dest)
        
        # Create classification subfolders and copy sub-crops
        classification_counts = {}
        for item in sub_items:
            if not os.path.exists(item['path']):
                continue
                
            classification = item.get('classified_class', 'unknown')
            r2_class = item.get('class', 'unknown')
            
            # Create classification subfolder
            class_folder = os.path.join(parent_folder, f"{classification}")
            os.makedirs(class_folder, exist_ok=True)
            
            # Generate unique filename
            classification_counts[classification] = classification_counts.get(classification, 0) + 1
            count = classification_counts[classification]
            
            filename = f"{count:02d}_{r2_class}_{os.path.basename(item['path'])}"
            dest_path = os.path.join(class_folder, filename)
            shutil.copy2(item['path'], dest_path)
        
        exported_folders.append(parent_folder)
        print(f"  Exported: {parent_name}/ with {len(sub_items)} sub-crops")
    
    # Create a summary text file
    summary_file = os.path.join(document_export_dir, "README.txt")
    with open(summary_file, 'w') as f:
        f.write(f"Image Export for: {original_pdf_name}\n")
        f.write("=" * 50 + "\n\n")
        f.write("Folder Structure:\n")
        f.write("- Each folder represents one detected section (R1 crop)\n")
        f.write("- 00_parent_* files are the original detected sections\n")
        f.write("- Subfolders are organized by R3 classification\n")
        f.write("- Files are numbered and include R2 detection type\n\n")
        f.write(f"Total parent sections: {len(parent_groups)}\n")
        f.write(f"Total sub-crops: {len(input_data)}\n\n")
        f.write("Classification Summary:\n")
        
        # Count classifications
        class_counts = {}
        for item in input_data:
            classification = item.get('classified_class', 'unknown')
            class_counts[classification] = class_counts.get(classification, 0) + 1
        
        for classification, count in sorted(class_counts.items()):
            f.write(f"  {classification}: {count} items\n")
    
    return document_export_dir

def create_zip_archive(folder_path, output_name):
    """Create a zip archive of the organized images"""
    if not folder_path or not os.path.exists(folder_path):
        return None
    
    import tempfile
    
    # Create zip in temp location first to avoid triggering Flask reload
    temp_dir = tempfile.gettempdir()
    temp_zip_path = os.path.join(temp_dir, f"{output_name}_images.zip")
    
    with zipfile.ZipFile(temp_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, folder_path)
                zipf.write(file_path, arcname)
    
    # Now move to final location after Excel processing is complete
    os.makedirs('outputs', exist_ok=True)
    final_zip_path = os.path.join('outputs', f"{output_name}_images.zip")
    shutil.move(temp_zip_path, final_zip_path)
    
    print(f"Created zip archive: {final_zip_path}")
    return final_zip_path

def get_results_dir(item):
    """Get the results directory for this item"""
    if not item or 'original_path' not in item:
        print("Warning: Invalid item or missing original_path")
        return os.path.join('results', 'default')
        
    try:
        # Get original file name from the path
        path_parts = item['original_path'].split(os.sep)
        if 'r1_crops' in path_parts:
            base_name = path_parts[path_parts.index('r1_crops') - 1]
        else:
            # Fallback to getting base name from the original path
            base_name = os.path.splitext(os.path.basename(item['original_path']))[0].split('_')[0]
        
        results_dir = os.path.join('results', base_name)
        os.makedirs(results_dir, exist_ok=True)  # Ensure directory exists
        return results_dir
    except Exception as e:
        print(f"Error getting results directory for {item.get('original_path', 'unknown')}: {str(e)}")
        fallback_dir = os.path.join('results', 'default')
        os.makedirs(fallback_dir, exist_ok=True)
        return fallback_dir

def process_r4(input_data):
    """
    Stage 4: Create Excel files with extraction results, with separate worksheets per subfolder.
    Each worksheet contains embedded images with their extraction results.
    Also creates a summary sheet in outputs/results.xlsx with all parent images and their location_position extracted texts.
    """
    print(f"R4 Processing started with {len(input_data) if input_data else 0} input items")
    
    if not input_data:
        print("Warning: No input data received for extraction")
        return []

    # Debug input data structure
    print("Input data sample:")
    for i, item in enumerate(input_data[:3]):  # Show first 3 items
        print(f"  Item {i}: {item}")
    
    workbook_paths = []
    summary_rows = []  # For the summary sheet
    zip_paths = []  # For the image zip files

    try:
        # Group data by R3 classification results instead of R2 folders
        classification_groups = {}
        for item in input_data:
            if not isinstance(item, dict) or 'path' not in item:
                print(f"Warning: Invalid item format: {item}")
                continue
            try:
                # Use R3 classification as the grouping key
                classification = item.get('classified_class', 'unknown')
                if classification not in classification_groups:
                    classification_groups[classification] = []
                classification_groups[classification].append(item)
            except Exception as e:
                print(f"Error processing item {item}: {str(e)}")
                continue
        
        if not classification_groups:
            print("Warning: No valid classification groups found")
            return []
        
        print(f"R4 Processing: Found {len(classification_groups)} classification groups:")
        for class_name, items in classification_groups.items():
            print(f"  {class_name}: {len(items)} items")
        
        # Determine base_name for the output file
        first_item = next(iter(classification_groups.values()))[0]
        original_pdf_name_for_file = "document" # Default

        # Try to get the original PDF name from the item structure
        # Assumed that R1 or R2 might add a 'source_pdf_name' or similar key,
        # or it can be derived from 'original_path' which should be the path to the R1 crop.
        source_pdf_name_found = False
        # Priority 1: Check for a specific key like 'source_pdf_filename' (adjust if R1/R2 uses a different key)
        if 'source_pdf_filename' in first_item and first_item['source_pdf_filename']:
            original_pdf_name_for_file = os.path.splitext(first_item['source_pdf_filename'])[0]
            source_pdf_name_found = True
        elif 'original_path' in first_item and first_item['original_path']:
            # original_path is expected to be like 'results/PDF_BASE_NAME/r1_crops/page_crop.png'
            # or directly a PDF file path if R1 was skipped or modified.
            path_obj = Path(first_item['original_path'])
            path_parts = path_obj.parts
            try:
                # Look for 'results' directory in the path
                results_index = path_parts.index('results')
                if len(path_parts) > results_index + 1:
                    # The part after 'results' is assumed to be the PDF base name folder
                    original_pdf_name_for_file = path_parts[results_index + 1]
                    source_pdf_name_found = True
            except ValueError:
                # 'results' not in path. This might mean original_path is the direct PDF path or an unexpected structure.
                # If it ends with .pdf, use its stem.
                if path_obj.suffix.lower() == '.pdf':
                    original_pdf_name_for_file = path_obj.stem
                    source_pdf_name_found = True
                else:
                    # Fallback for other image files, try to get a meaningful name if not a generic crop name
                    temp_name = path_obj.stem
                    if not re.match(r"page\d+_crop\d+(?:_\d+)?", temp_name): # Adjusted regex for names like page0_crop0_0
                        original_pdf_name_for_file = temp_name
                        source_pdf_name_found = True
        
        if not source_pdf_name_found:
            # If still not found, iterate through all items to find a usable original_path
            for item_list in classification_groups.values():
                for item_from_list in item_list:
                    if 'original_path' in item_from_list and item_from_list['original_path']:
                        path_obj_iter = Path(item_from_list['original_path'])
                        path_parts_iter = path_obj_iter.parts
                        try:
                            results_index_iter = path_parts_iter.index('results')
                            if len(path_parts_iter) > results_index_iter + 1:
                                original_pdf_name_for_file = path_parts_iter[results_index_iter + 1]
                                source_pdf_name_found = True
                                break
                        except ValueError:
                            if path_obj_iter.suffix.lower() == '.pdf':
                                original_pdf_name_for_file = path_obj_iter.stem
                                source_pdf_name_found = True
                                break
                if source_pdf_name_found:
                    break
            if not source_pdf_name_found:
                 print(f"Warning: Could not reliably determine original PDF name from any item. Using default '{original_pdf_name_for_file}'.")

        # Create one Excel file with multiple worksheets FIRST
        try:
            print("Starting Excel file creation...")
            outputs_dir = 'outputs'
            os.makedirs(outputs_dir, exist_ok=True)
            # Sanitize the pdf name for use in filename
            safe_pdf_name = re.sub(r'[^a-zA-Z0-9_\-.]', '_', original_pdf_name_for_file) # Allow dots for extensions if any part of name has it
            excel_path = os.path.join(outputs_dir, f'{safe_pdf_name}_extraction_results.xlsx')
            
            print(f"Creating Excel file: {excel_path}")
            
            workbook = Workbook()
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook.active)
                
            print(f"Processing {len(classification_groups)} classification groups for worksheets...")
            
            # Process each classification group in a separate worksheet
            worksheets_created = 0
            for classification_name, items in classification_groups.items():
                try:
                    print(f"  Creating worksheet for classification: {classification_name} ({len(items)} items)")
                    
                    # Skip empty classifications
                    if not items:
                        print(f"    Skipping empty classification: {classification_name}")
                        continue
                    
                    worksheet = workbook.create_sheet(title=classification_name)
                    parent_groups = {}
                    for item in items:
                        if not all(key in item for key in ['original_path', 'path']):
                            continue
                        parent_path = item['original_path']
                        if parent_path not in parent_groups:
                            parent_groups[parent_path] = []
                        parent_groups[parent_path].append(item)
                    
                    # Skip if no valid parent groups
                    if not parent_groups:
                        print(f"    No valid parent groups for classification: {classification_name}")
                        workbook.remove(worksheet)
                        continue
                    worksheet.column_dimensions['A'].width = 20  # Type
                    worksheet.column_dimensions['B'].width = 40  # Image
                    worksheet.column_dimensions['C'].width = 30  # Classification  
                    worksheet.column_dimensions['D'].width = 60  # Extracted Text
                    row_height = 150
                    headers = ['Type', 'Image', 'Classification', 'Extracted Text']
                    for col, header in enumerate(headers, 1):
                        cell = worksheet.cell(row=1, column=col)
                        cell.value = header
                        cell.font = cell.font.copy(bold=True)
                    row_idx = 1
                    for parent_path, sub_items in parent_groups.items():
                        if not os.path.exists(parent_path):
                            continue
                        try:
                            # Add parent image at the top
                            img = Image.open(parent_path)
                            img.thumbnail((300, 300))
                            img_byte_arr = BytesIO()
                            img.save(img_byte_arr, format='PNG')
                            img_byte_arr.seek(0)
                            xl_img = XLImage(img_byte_arr)
                            xl_img.anchor = f'B{row_idx + 1}'
                            worksheet.add_image(xl_img)
                            worksheet.row_dimensions[row_idx + 1].height = row_height
                            worksheet.cell(row=row_idx + 1, column=1).value = 'Parent Image'
                            worksheet.cell(row=row_idx + 1, column=3).value = ''
                            parent_text = text_extractor.extract_text(parent_path)
                            worksheet.cell(row=row_idx + 1, column=4).value = parent_text
                            # For summary: add parent image and all location_position extracted texts
                            summary_rows.append({'parent_path': parent_path, 'parent_img_bytes': img_byte_arr.getvalue(), 'location_texts': []})
                            summary_idx = len(summary_rows) - 1
                            row_idx += 1
                            for sub_item in sub_items:
                                if not os.path.exists(sub_item['path']):
                                    continue
                                try:
                                    sub_img = Image.open(sub_item['path'])
                                    sub_img.thumbnail((300, 300))
                                    sub_img_byte_arr = BytesIO()
                                    sub_img.save(sub_img_byte_arr, format='PNG')
                                    sub_img_byte_arr.seek(0)
                                    sub_xl_img = XLImage(sub_img_byte_arr)
                                    sub_xl_img.anchor = f'B{row_idx + 1}'
                                    worksheet.add_image(sub_xl_img)
                                    worksheet.row_dimensions[row_idx + 1].height = row_height
                                    worksheet.cell(row=row_idx + 1, column=1).value = 'Sub-crop'
                                    # Show both R2 class (from folder structure) and R3 classification
                                    r2_class = sub_item.get('class', 'Unknown')
                                    r3_class = sub_item.get('classified_class', 'Unknown')
                                    worksheet.cell(row=row_idx + 1, column=3).value = f"R2: {r2_class} | R3: {r3_class}"
                                    sub_text = text_extractor.extract_text(sub_item['path'])
                                    # If location_position, add to summary
                                    if r3_class and 'location' in r3_class.lower():
                                        summary_rows[summary_idx]['location_texts'].append(sub_text)
                                    worksheet.cell(row=row_idx + 1, column=4).value = sub_text
                                    row_idx += 1
                                except Exception as e:
                                    print(f"Error processing sub-crop image: {str(e)}")
                                    continue
                        except Exception as e:
                            print(f"Error processing parent image: {str(e)}")
                            continue
                    
                    worksheets_created += 1
                    print(f"    Completed worksheet: {classification_name}")
                    
                except Exception as e:
                    print(f"Error processing worksheet {classification_name}: {str(e)}")
                    continue
            
            print(f"Total worksheets created: {worksheets_created}")
            
            # Don't save if no worksheets were created
            if worksheets_created == 0:
                print("Warning: No worksheets were created. Skipping Excel file creation.")
            else:
                # Save main workbook
                print(f"Saving Excel workbook to: {excel_path}")
                workbook.save(excel_path)
                workbook_paths.append(excel_path)
                print(f"Successfully created Excel file: {excel_path} (Size: {os.path.getsize(excel_path)} bytes)")
            
            # Create summary workbook in outputs/results.xlsx
            print("Creating summary workbook...")
            summary_path = os.path.join('outputs', 'results.xlsx')
            os.makedirs('outputs', exist_ok=True)
            summary_wb = Workbook()
            summary_ws = summary_wb.active
            summary_ws.title = 'Summary'
            summary_ws.column_dimensions['A'].width = 40
            summary_ws.column_dimensions['B'].width = 80
            summary_ws.column_dimensions['C'].width = 80
            summary_ws.cell(row=1, column=1).value = 'Parent Image'
            summary_ws.cell(row=1, column=2).value = 'Location Position Extracted Texts'
            for idx, row in enumerate(summary_rows, 2):
                # Add parent image
                img_byte_arr = BytesIO(row['parent_img_bytes'])
                xl_img = XLImage(img_byte_arr)
                xl_img.anchor = f'A{idx}'
                summary_ws.add_image(xl_img)
                # Add all location_position extracted texts
                summary_ws.cell(row=idx, column=2).value = '\n'.join(row['location_texts'])
            summary_wb.save(summary_path)
            workbook_paths.append(summary_path)
            print(f"Summary saved to {summary_path} (Size: {os.path.getsize(summary_path)} bytes)")
        except Exception as e:
            print(f"Error creating Excel file: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # Create organized image export for verification AFTER Excel is complete
        print("Creating organized image export...")
        export_folder = create_organized_image_export(input_data, original_pdf_name_for_file)
        if export_folder:
            # Create zip archive of the images
            zip_path = create_zip_archive(export_folder, original_pdf_name_for_file)
            if zip_path:
                zip_paths.append(zip_path)
        
        # Debug output
        print(f"R4 Processing complete:")
        print(f"  Excel files created: {len(workbook_paths)}")
        print(f"  Image zip files created: {len(zip_paths)}")
        for path in workbook_paths:
            print(f"    Excel: {path}")
        for path in zip_paths:
            print(f"    Zip: {path}")
        
        # Return both workbook paths and zip paths
        all_output_paths = workbook_paths + zip_paths
        print(f"  Total output files: {len(all_output_paths)}")
        return all_output_paths
    except Exception as e:
        print(f"Error in process_r4: {str(e)}")
        return []
