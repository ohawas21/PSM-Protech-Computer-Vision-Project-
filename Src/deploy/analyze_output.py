#!/usr/bin/env python3
"""
Script to analyze the current output format and show what it produces
"""
import os
import pandas as pd
from openpyxl import load_workbook
import yaml

def analyze_excel_output():
    """Analyze the current Excel outputs"""
    print("=" * 60)
    print("OUTPUT FORMAT ANALYSIS")
    print("=" * 60)
    
    # Check outputs directory
    outputs_dir = 'outputs'
    if not os.path.exists(outputs_dir):
        print(f"Outputs directory '{outputs_dir}' not found")
        return
    
    excel_files = [f for f in os.listdir(outputs_dir) if f.endswith('.xlsx')]
    print(f"Found {len(excel_files)} Excel files in outputs:")
    
    for excel_file in excel_files:
        file_path = os.path.join(outputs_dir, excel_file)
        print(f"\n📄 {excel_file}")
        print(f"   Size: {os.path.getsize(file_path)} bytes")
        
        try:
            # Load workbook to analyze structure
            wb = load_workbook(file_path)
            print(f"   Worksheets ({len(wb.sheetnames)}):")
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                max_row = ws.max_row
                max_col = ws.max_column
                print(f"     • {sheet_name}: {max_row} rows × {max_col} columns")
                
                # Show headers if available
                if max_row > 0:
                    headers = []
                    for col in range(1, min(max_col + 1, 5)):  # First 4 columns
                        cell_value = ws.cell(row=1, column=col).value
                        if cell_value:
                            headers.append(str(cell_value))
                    if headers:
                        print(f"       Headers: {' | '.join(headers)}")
                
                # Count images
                image_count = len(ws._images) if hasattr(ws, '_images') else 0
                print(f"       Images embedded: {image_count}")
                
        except Exception as e:
            print(f"   Error reading Excel file: {str(e)}")

def analyze_results_structure():
    """Analyze the results directory structure"""
    print("\n" + "=" * 60)
    print("RESULTS DIRECTORY STRUCTURE")
    print("=" * 60)
    
    results_dir = 'results'
    if not os.path.exists(results_dir):
        print(f"Results directory '{results_dir}' not found")
        return
    
    subdirs = [d for d in os.listdir(results_dir) if os.path.isdir(os.path.join(results_dir, d))]
    print(f"Found {len(subdirs)} result subdirectories:")
    
    for subdir in subdirs:
        subdir_path = os.path.join(results_dir, subdir)
        print(f"\n📁 {subdir}/")
        
        # List contents
        for item in os.listdir(subdir_path):
            item_path = os.path.join(subdir_path, item)
            if os.path.isdir(item_path):
                sub_items = os.listdir(item_path)
                print(f"   📁 {item}/ ({len(sub_items)} items)")
                
                # Show first few items as examples
                for i, sub_item in enumerate(sub_items[:3]):
                    sub_item_path = os.path.join(item_path, sub_item)
                    if os.path.isfile(sub_item_path):
                        size = os.path.getsize(sub_item_path)
                        print(f"      📄 {sub_item} ({size} bytes)")
                    elif os.path.isdir(sub_item_path):
                        sub_sub_items = len(os.listdir(sub_item_path))
                        print(f"      📁 {sub_item}/ ({sub_sub_items} items)")
                
                if len(sub_items) > 3:
                    print(f"      ... and {len(sub_items) - 3} more items")
                    
            else:
                size = os.path.getsize(item_path)
                print(f"   📄 {item} ({size} bytes)")

def analyze_config():
    """Analyze current configuration"""
    print("\n" + "=" * 60)
    print("CURRENT CONFIGURATION")
    print("=" * 60)
    
    config_path = 'config.yaml'
    if os.path.exists(config_path):
        with open(config_path, 'r') as f:
            config = yaml.safe_load(f)
        
        print("Configuration settings:")
        for key, value in config.items():
            print(f"  {key}: {value}")
    else:
        print("Config file not found")

def show_processing_flow():
    """Show the processing flow and what each stage produces"""
    print("\n" + "=" * 60)
    print("PROCESSING FLOW EXPLANATION")
    print("=" * 60)
    
    flow = [
        ("Stage 1 (R1)", "PDF/Image → YOLO Detection → Individual Crops", "results/[DOC]/r1_crops/"),
        ("Stage 2 (R2)", "R1 Crops → YOLO Classification → Categorized Sub-crops", "results/[DOC]/r2_crops/"),
        ("Stage 3 (R3)", "R2 Sub-crops → Classification → Labeled Data", "In-memory data with 'classified_class'"),
        ("Stage 4 (R4)", "Classified Data → Excel Reports → Text Extraction", "outputs/[DOC]_extraction_results.xlsx")
    ]
    
    for stage, description, output in flow:
        print(f"\n🔄 {stage}")
        print(f"   Process: {description}")
        print(f"   Output: {output}")
    
    print(f"\n📊 EXCEL OUTPUT FORMAT:")
    print(f"   • One workbook per processed document")
    print(f"   • One worksheet per R3 classification category")
    print(f"   • Each worksheet contains:")
    print(f"     - Parent images (R1 crops) with OCR text")
    print(f"     - Sub-crop images (R2 outputs) with R2+R3 classifications")
    print(f"     - Extracted text from each image")
    print(f"     - All images embedded directly in Excel")
    print(f"   • Summary file with location-based extractions")

if __name__ == "__main__":
    analyze_config()
    analyze_results_structure()
    analyze_excel_output()
    show_processing_flow()
    
    print(f"\n" + "=" * 60)
    print("ANALYSIS COMPLETE")
    print("=" * 60)
