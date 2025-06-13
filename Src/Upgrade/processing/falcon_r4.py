import os
import re
import logging
from pathlib import Path
from io import BytesIO
import warnings
# import pandas as pd # Not used directly in this version
import camelot
import cv2
import numpy as np
from PIL import Image
from pdf2image import convert_from_bytes
import pytesseract
# Ensure openpyxl is imported at the top level
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger(__name__)

# Suppress Camelot warnings for image-based pages
warnings.filterwarnings('ignore', category=UserWarning)

# Tesseract config: only digits and dot
TESS_CONFIG = "--psm 7 --oem 3 -c tessedit_char_whitelist=0123456789."

class TextExtractor:
    def __init__(self):
        self.temp_dir = Path("temp_processing")
        self.temp_dir.mkdir(parents=True, exist_ok=True)

    def convert_to_pdf(self, image_path: str) -> Path:
        """Convert image to PDF for processing"""
        buf = BytesIO()
        Image.open(image_path).convert("RGB").save(buf, format="PDF")
        buf.seek(0)
        pdf_path = self.temp_dir / f"{Path(image_path).stem}.pdf"
        pdf_path.write_bytes(buf.getvalue())
        return pdf_path

    def try_camelot(self, pdf_path: Path) -> str:
        """Try to extract text using Camelot"""
        try:
            tables = camelot.read_pdf(str(pdf_path), flavor="stream", pages="1")
            if tables:
                text = "\n".join([table.df.to_string() for table in tables])
                logger.info(f"Successfully extracted text using Camelot from {pdf_path.name}")
                return text
        except Exception as e:
            logger.warning(f"Camelot extraction failed for {pdf_path.name}: {str(e)}")
        return ""

    def render_to_image(self, file_path: str) -> Image.Image:
        """Convert PDF to image or load image directly"""
        try:
            if file_path.lower().endswith('.pdf'):
                pages = convert_from_bytes(Path(file_path).read_bytes(), dpi=300)
                return pages[0] if pages else None
            else:
                return Image.open(file_path)
        except Exception as e:
            logger.error(f"Error converting {file_path}: {str(e)}")
            return None

    def preprocess_image(self, img: Image.Image) -> Image.Image:
        """Preprocess image for better OCR"""
        arr = cv2.cvtColor(np.array(img), cv2.COLOR_RGB2GRAY)
        blur = cv2.GaussianBlur(arr, (3,3), 0)
        _, th = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        inv = cv2.bitwise_not(th)
        up = cv2.resize(inv, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
        return Image.fromarray(up)

    def extract_text(self, file_path: str) -> str:
        """Extract text from file using multiple methods"""
        file_path = str(file_path)
        text = ""
        
        # Try Camelot first for PDFs
        if file_path.lower().endswith('.pdf'):
            text = self.try_camelot(Path(file_path))
            if text.strip():
                return text

        # If Camelot fails or it's an image, try OCR
        img = self.render_to_image(file_path)
        if img:
            preprocessed = self.preprocess_image(img)
            text = pytesseract.image_to_string(preprocessed)
            logger.info(f"Extracted text using OCR from {Path(file_path).name}")
            return text.strip()
            
        return ""

# Global extractor instance
text_extractor = TextExtractor()

def get_results_dir(item):
    """Get the results directory for this item"""
    if not item or 'original_path' not in item:
        print("Warning: Invalid item or missing original_path")
        return os.path.join('results', 'default')
        
    try:
        # Get original file name from the path
        path_parts = item['original_path'].split(os.sep)
        if 'r1_crops' in path_parts:
            base_name = path_parts[path_parts.index('r1_crops') - 1]
        else:
            # Fallback to getting base name from the original path
            base_name = os.path.splitext(os.path.basename(item['original_path']))[0].split('_')[0]
        
        results_dir = os.path.join('results', base_name)
        os.makedirs(results_dir, exist_ok=True)  # Ensure directory exists
        return results_dir
    except Exception as e:
        print(f"Error getting results directory for {item.get('original_path', 'unknown')}: {str(e)}")
        fallback_dir = os.path.join('results', 'default')
        os.makedirs(fallback_dir, exist_ok=True)
        return fallback_dir

def process_r4(input_data):
    """
    Stage 4: Create Excel files with extraction results, with separate worksheets per subfolder.
    Each worksheet contains embedded images with their extraction results.
    Also creates a summary sheet in outputs/results.xlsx with all parent images and their location_position extracted texts.
    """
    if not input_data:
        print("Warning: No input data received for extraction")
        return []

    workbook_paths = []
    summary_rows = []  # For the summary sheet

    try:
        # Group data by source folders (from R2)
        folder_groups = {}
        for item in input_data:
            if not isinstance(item, dict) or 'path' not in item:
                print(f"Warning: Invalid item format: {item}")
                continue
            try:
                parent_folder = os.path.basename(os.path.dirname(item['path']))
                if not parent_folder:
                    print(f"Warning: Could not determine parent folder for {item['path']}")
                    continue
                if parent_folder not in folder_groups:
                    folder_groups[parent_folder] = []
                folder_groups[parent_folder].append(item)
            except Exception as e:
                print(f"Error processing item {item}: {str(e)}")
                continue
        if not folder_groups:
            print("Warning: No valid folder groups found")
            return []
        
        # Determine base_name for the output file
        first_item = next(iter(folder_groups.values()))[0]
        original_pdf_name_for_file = "document" # Default

        # Try to get the original PDF name from the item structure
        # Assumed that R1 or R2 might add a 'source_pdf_name' or similar key,
        # or it can be derived from 'original_path' which should be the path to the R1 crop.
        source_pdf_name_found = False
        # Priority 1: Check for a specific key like 'source_pdf_filename' (adjust if R1/R2 uses a different key)
        if 'source_pdf_filename' in first_item and first_item['source_pdf_filename']:
            original_pdf_name_for_file = os.path.splitext(first_item['source_pdf_filename'])[0]
            source_pdf_name_found = True
        elif 'original_path' in first_item and first_item['original_path']:
            # original_path is expected to be like 'results/PDF_BASE_NAME/r1_crops/page_crop.png'
            # or directly a PDF file path if R1 was skipped or modified.
            path_obj = Path(first_item['original_path'])
            path_parts = path_obj.parts
            try:
                # Look for 'results' directory in the path
                results_index = path_parts.index('results')
                if len(path_parts) > results_index + 1:
                    # The part after 'results' is assumed to be the PDF base name folder
                    original_pdf_name_for_file = path_parts[results_index + 1]
                    source_pdf_name_found = True
            except ValueError:
                # 'results' not in path. This might mean original_path is the direct PDF path or an unexpected structure.
                # If it ends with .pdf, use its stem.
                if path_obj.suffix.lower() == '.pdf':
                    original_pdf_name_for_file = path_obj.stem
                    source_pdf_name_found = True
                else:
                    # Fallback for other image files, try to get a meaningful name if not a generic crop name
                    temp_name = path_obj.stem
                    if not re.match(r"page\d+_crop\d+(?:_\d+)?", temp_name): # Adjusted regex for names like page0_crop0_0
                        original_pdf_name_for_file = temp_name
                        source_pdf_name_found = True
        
        if not source_pdf_name_found:
            # If still not found, iterate through all items to find a usable original_path
            for item_list in folder_groups.values():
                for item_from_list in item_list:
                    if 'original_path' in item_from_list and item_from_list['original_path']:
                        path_obj_iter = Path(item_from_list['original_path'])
                        path_parts_iter = path_obj_iter.parts
                        try:
                            results_index_iter = path_parts_iter.index('results')
                            if len(path_parts_iter) > results_index_iter + 1:
                                original_pdf_name_for_file = path_parts_iter[results_index_iter + 1]
                                source_pdf_name_found = True
                                break
                        except ValueError:
                            if path_obj_iter.suffix.lower() == '.pdf':
                                original_pdf_name_for_file = path_obj_iter.stem
                                source_pdf_name_found = True
                                break
                if source_pdf_name_found:
                    break
            if not source_pdf_name_found:
                 print(f"Warning: Could not reliably determine original PDF name from any item. Using default '{original_pdf_name_for_file}'.")


        # Create one Excel file with multiple worksheets
        try:
            outputs_dir = 'outputs'
            os.makedirs(outputs_dir, exist_ok=True)
            # Sanitize the pdf name for use in filename
            safe_pdf_name = re.sub(r'[^a-zA-Z0-9_\-.]', '_', original_pdf_name_for_file) # Allow dots for extensions if any part of name has it
            excel_path = os.path.join(outputs_dir, f'{safe_pdf_name}_extraction_results.xlsx')
            
            workbook = Workbook()
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook.active)
            # Process each folder in a separate worksheet
            for folder_name, items in folder_groups.items():
                try:
                    worksheet = workbook.create_sheet(title=folder_name)
                    parent_groups = {}
                    for item in items:
                        if not all(key in item for key in ['original_path', 'path']):
                            continue
                        parent_path = item['original_path']
                        if parent_path not in parent_groups:
                            parent_groups[parent_path] = []
                        parent_groups[parent_path].append(item)
                    worksheet.column_dimensions['A'].width = 20
                    worksheet.column_dimensions['B'].width = 40
                    worksheet.column_dimensions['C'].width = 20
                    worksheet.column_dimensions['D'].width = 50
                    row_height = 150
                    headers = ['Type', 'Image', 'Classification', 'Falcon 4 Extraction']
                    for col, header in enumerate(headers, 1):
                        cell = worksheet.cell(row=1, column=col)
                        cell.value = header
                        cell.font = cell.font.copy(bold=True)
                    row_idx = 1
                    for parent_path, sub_items in parent_groups.items():
                        if not os.path.exists(parent_path):
                            continue
                        try:
                            # Add parent image at the top
                            img = Image.open(parent_path)
                            img.thumbnail((300, 300))
                            img_byte_arr = BytesIO()
                            img.save(img_byte_arr, format='PNG')
                            img_byte_arr.seek(0)
                            xl_img = XLImage(img_byte_arr)
                            xl_img.anchor = f'B{row_idx + 1}'
                            worksheet.add_image(xl_img)
                            worksheet.row_dimensions[row_idx + 1].height = row_height
                            worksheet.cell(row=row_idx + 1, column=1).value = 'Parent Image'
                            worksheet.cell(row=row_idx + 1, column=3).value = ''
                            parent_text = text_extractor.extract_text(parent_path)
                            worksheet.cell(row=row_idx + 1, column=4).value = parent_text
                            # For summary: add parent image and all location_position extracted texts
                            summary_rows.append({'parent_path': parent_path, 'parent_img_bytes': img_byte_arr.getvalue(), 'location_texts': []})
                            summary_idx = len(summary_rows) - 1
                            row_idx += 1
                            for sub_item in sub_items:
                                if not os.path.exists(sub_item['path']):
                                    continue
                                try:
                                    sub_img = Image.open(sub_item['path'])
                                    sub_img.thumbnail((300, 300))
                                    sub_img_byte_arr = BytesIO()
                                    sub_img.save(sub_img_byte_arr, format='PNG')
                                    sub_img_byte_arr.seek(0)
                                    sub_xl_img = XLImage(sub_img_byte_arr)
                                    sub_xl_img.anchor = f'B{row_idx + 1}'
                                    worksheet.add_image(sub_xl_img)
                                    worksheet.row_dimensions[row_idx + 1].height = row_height
                                    worksheet.cell(row=row_idx + 1, column=1).value = 'Sub-crop'
                                    worksheet.cell(row=row_idx + 1, column=3).value = sub_item.get('classified_class', 'Unknown')
                                    sub_text = text_extractor.extract_text(sub_item['path'])
                                    # If location_position, add to summary
                                    if sub_item.get('classified_class', '').startswith('location_position'):
                                        summary_rows[summary_idx]['location_texts'].append(sub_text)
                                    worksheet.cell(row=row_idx + 1, column=4).value = f"{sub_item.get('classified_class', '')}: {sub_text}"
                                    row_idx += 1
                                except Exception as e:
                                    print(f"Error processing sub-crop image: {str(e)}")
                                    continue
                        except Exception as e:
                            print(f"Error processing parent image: {str(e)}")
                            continue
                except Exception as e:
                    print(f"Error processing worksheet {folder_name}: {str(e)}")
                    continue
            workbook.save(excel_path)
            workbook_paths.append(excel_path)
            print(f"Successfully created Excel file with multiple worksheets and embedded images: {excel_path}")
            # Create summary workbook in outputs/results.xlsx
            summary_path = os.path.join('outputs', 'results.xlsx')
            os.makedirs('outputs', exist_ok=True)
            summary_wb = Workbook()
            summary_ws = summary_wb.active
            summary_ws.title = 'Summary'
            summary_ws.column_dimensions['A'].width = 40
            summary_ws.column_dimensions['B'].width = 80
            summary_ws.column_dimensions['C'].width = 80
            summary_ws.cell(row=1, column=1).value = 'Parent Image'
            summary_ws.cell(row=1, column=2).value = 'Location Position Extracted Texts'
            for idx, row in enumerate(summary_rows, 2):
                # Add parent image
                img_byte_arr = BytesIO(row['parent_img_bytes'])
                xl_img = XLImage(img_byte_arr)
                xl_img.anchor = f'A{idx}'
                summary_ws.add_image(xl_img)
                # Add all location_position extracted texts
                summary_ws.cell(row=idx, column=2).value = '\n'.join(row['location_texts'])
            summary_wb.save(summary_path)
            print(f"Summary saved to {summary_path}")
        except Exception as e:
            print(f"Error creating Excel file: {str(e)}")
        return workbook_paths
    except Exception as e:
        print(f"Error in process_r4: {str(e)}")
        return []
